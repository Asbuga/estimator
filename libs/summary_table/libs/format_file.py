import os

import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font



# Open file.
output_file = "summary_table.xlsx"
wb = openpyxl.load_workbook(output_file)
ws = wb.worksheets[0]

# Settings summary table.
# Head row style.
head_row_style = Alignment(wrap_text=True, horizontal="center", vertical="top")
head_row_color = PatternFill(
    start_color="ffe599", end_color="ffe599", fill_type="solid")

border_style = Border(
    left=Side(border_style="thin", color='000000'),
    right=Side(border_style="thin", color='000000'),
    top=Side(border_style="thin", color='000000'),
    bottom=Side(border_style="thin", color='000000')
)


# We set the sizes of the cells.
# Header row.
ws.row_dimensions[1].height = 100

# All cells.
for row in ws.iter_rows():
    for cell in row:

        # For all cells.
        cell.alignment = Alignment(vertical="top")
        cell.border = border_style

        # Head row.
        if cell.row == 1:
            cell.fill = head_row_color
            cell.alignment = head_row_style
            cell.font = Font(bold=True)

            if cell.column_letter != "J":
                ws.column_dimensions[cell.column_letter].width = 10
            else:
                ws.column_dimensions[cell.column_letter].width = 50

        elif cell.column_letter == "J":
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        elif cell.column_letter in ["A", "C", "D", "H", "K"]:
            cell.alignment = head_row_style        


# Закриття письменника Excel
wb.save(output_file)

os.system("start excel summary_table.xlsx")
